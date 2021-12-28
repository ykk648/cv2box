# -- coding: utf-8 --
# @Time : 2021/12/20
# @Author : ykk648
# @Project : https://github.com/ykk648/cv2box
from pathlib import Path
import openpyxl.drawing.image as openpyxl_image
from openpyxl import load_workbook, Workbook


class CVExcel:
    def __init__(self, xlsx_path):

        if Path(xlsx_path).exists():
            self.wb = load_workbook(xlsx_path)
        else:
            self.wb = Workbook()
        self.sheet = self.wb.active

    @property
    def create_sheet(self):
        self.sheet = self.wb.create_sheet(title="Data")
        return self.sheet

    def mod_size(self, location, size=(7, 7)):
        self.sheet.column_dimensions[location[0]].width = size[0]
        self.sheet.row_dimensions[int(location[1:])].height = size[1]

    def insert_image(self, insert_location, image_path, image_new_size=None):

        img = openpyxl_image.Image(image_path)

        if image_new_size:
            img.width, img.height = image_new_size

        self.sheet[insert_location] = ""
        self.sheet.add_image(img, insert_location)

    def insert_words(self, insert_location, words_):
        self.sheet[insert_location] = words_

    def save(self, new_xlsx_path):
        self.wb.save(new_xlsx_path)
